from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import struct
import matplotlib.pyplot as plt
import numpy as np
#Open Image and destination XLS file
img = plt.imread("pic.tif")
dest_filename = 'empty_book.xlsx'
rows, cols, rgb = np.shape(img)
R, G, B = img[:,:,0], img[:,:,1], img[:,:,2]
wb = Workbook()
ws = wb.active
# Set the width and height of cells (1 width and 5.5 height is square to replicate pixel)
for i in range(len(R)):
    ws.column_dimensions[get_column_letter(i+1)].width = 1
    ws.row_dimensions[i+1].height = 5.5
# Traverse through the metrix of pixes. Extract Red, Green and Blue value and set the fill the cell at i, j location with that color.
for i in range(len(R)):
    for j in range(len(R[i])):
        R_comp = R[i][j]
        G_comp = G[i][j]
        B_comp = B[i][j]
        rgb = (R_comp,G_comp,B_comp)
        final = bytes.hex(struct.pack('BBB',*rgb))
        cell_id = get_column_letter(j+1)+str(i+1)
        ws[cell_id].fill = PatternFill("solid", start_color=final)
wb.save(filename = dest_filename)
